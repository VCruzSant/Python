# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# carregando um arquivo do excel
workbook = load_workbook(WORKBOOK_PATH)

# nome da página pra planilha:
sheet_name = 'My Sheet'

# dizendo pro python que quero mexer na planilha que eu selecionei
worksheet: Worksheet = workbook[sheet_name]  # type: ignore

for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

# editando valor de uma celular em especifico:
worksheet['B3']. value = 14


# salvando a planilha configurada no caminho desejado
workbook.save(WORKBOOK_PATH)
